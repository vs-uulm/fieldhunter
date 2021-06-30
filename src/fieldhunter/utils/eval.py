import os, csv, logging
from typing import Any

from openpyxl import Workbook, utils

from nemere.inference.segments import MessageSegment
from nemere.validation.dissectorMatcher import MessageComparator
from openpyxl.worksheet.worksheet import Worksheet

from fieldhunter.inference.fieldtypes import *


def csvAppend(reportFolder: str, fileName: str, header: List[str], rows: Iterable[Iterable[Any]]):
    csvpath = os.path.join(reportFolder, fileName + '.csv')
    csvWriteHead = False if os.path.exists(csvpath) else True

    print('Write statistics to {}...'.format(csvpath))
    with open(csvpath, 'a') as csvfile:
        statisticscsv = csv.writer(csvfile)
        if csvWriteHead:
            statisticscsv.writerow(header)
        statisticscsv.writerows(rows)


class FieldTypeReport(object):

    headers = ["hexbytes", "segment offset", "segment end",
               "overlap ratio", "overlap index", "overlap offset", "overlap end", "overlap value",
               "message date", "message type", "field name", "field type", "TP/FP", "isVisible"]
    # (column isVisible could also be called: "not hidden by other type")

    overviewHeaders = [
        "field type", "FN", "FP", "TP", "P", "R"
    ]

    def __init__(self, fieldtype: FieldType, comparator: MessageComparator,
                 segmentedMessages: Dict[AbstractMessage, List[MessageSegment]] = None):
        """

        :param fieldtype: The field type object to generate a report for.
        :param comparator: A NEMERE MessageComparator to look up the true fields overlapping our inference.
        :param segmentedMessages: Optional Dict of segmented messages to check whether another field type got
            precedence for single inference instances. see fieldhunter.inference.fieldtypes#precedence and
            fieldhunter.inference.common#segmentedMessagesAndSymbols
        """
        self._fieldtype = fieldtype
        self._comparator = comparator
        self._segmentedMessages = segmentedMessages

    def lookupOverlap(self):
        """
        Lookup the overlap with the ground truth for all segments inferred for the given FieldHunter field type.

        :return: table (list of lists) of statistics for each inferred segment from field type, according to the
            columns given in FieldTypeReport#headers.
        """
        tabdata = list()

        for seg in (seg for msgsegs in self._fieldtype.segments for seg in msgsegs if msgsegs):
            # field: from ground true; seg(ment): inferred; overlap: intersection of field and segment
            overlapRatio, overlapIndex, overlapOffset, overlapEnd = self._comparator.fieldOverlap(seg)
            messagetype, fieldname, fieldtype = self._comparator.lookupField(seg)
            overlapValue = "'" + seg.message.data[overlapOffset:overlapEnd].hex() + "'"

            # determine what is a TP/FP using GroundTruth
            tpfp = fieldname in GroundTruth.fieldtypes[self.typelabel]

            # check the precedence of multiple overlapping inferred fields
            isVisible = seg in chain.from_iterable(self._segmentedMessages.values())\
                if self._segmentedMessages is not None else "n/a"

            tabdata.append(["'" + seg.bytes.hex() + "'", seg.offset, seg.nextOffset,
                            overlapRatio, overlapIndex, overlapOffset, overlapEnd, overlapValue,
                            seg.message.date, messagetype, fieldname, fieldtype, tpfp, isVisible])
        return tabdata

    @property
    def typelabel(self):
        """The label for the field type this report is generated for."""
        return self._fieldtype.typelabel

    def countTrueOccurrences(self):
        counter = 0
        for fieldname in GroundTruth.fieldtypes[self.typelabel]:
            counter += len(self._comparator.lookupValues4FieldName(fieldname))
        return counter

    def addXLworksheet(self, workbook: Workbook, overview: str=None):
        """Add data as worksheet to a openpyxl workbook. The caller needs to take take to write to file afterwards."""
        worksheet = workbook.create_sheet(self.typelabel)
        worksheet.append(FieldTypeReport.headers)
        for row in self.lookupOverlap():
            worksheet.append(row)
        if overview is not None:
            try:
                ovSheet = workbook[overview]  # type: Worksheet
                currentRow = ovSheet.max_row + 1
                tpCoord = f"{utils.get_column_letter(4)}{currentRow}"
                ovSheet.append([
                    self.typelabel,
                    f"={self.countTrueOccurrences()} - {tpCoord}", # "FN"
                    f"=COUNTIF({utils.quote_sheetname(self.typelabel)}!M:M,FALSE())", # "=FP"
                    f"=COUNTIF({utils.quote_sheetname(self.typelabel)}!M:M,TRUE())",  # "=TP"
                    f"=D{currentRow}/(D{currentRow}+C{currentRow})", # P
                    f"=D{currentRow}/(D{currentRow}+B{currentRow})", # R
                ])
            except KeyError:
                logging.getLogger(__name__).info("Overview sheet with title", overview, "not found. "
                                                 "Not writing overview.")
        return workbook


class GroundTruth(object):
    """tshark dissector field names for sample protocols mapped from the FieldHunter field type class."""
    fieldtypes = {
        MSGlen.typelabel:    ["nbss.length"],
        MSGtype.typelabel:   ["dhcp.option.dhcp", "ntp.flags", "ntp.stratum", "dns.flags",
                              "nbns.flags", "smb.cmd", "smb.flags", ],
        HostID.typelabel:    ["dhcp.ip.client", "dhcp.ip.your", "dhcp.ip.server", "ntp.refid"],
        SessionID.typelabel: ["dhcp.id", "smb.pid", "smb.uid", "smb.mid"],
        TransID.typelabel:   ["dns.id", "nbns.id"],
        Accumulator.typelabel: []
    }

    def __init__(self, comparator:MessageComparator, endianness: str = "big"):
        self._comparator = comparator
        self._endianness = endianness
        logging.getLogger(__name__).setLevel(logging.DEBUG)

    def entropyPerField(self, fieldname: str):
        """Collect true fields values and calculate their entropy for the current trace."""
        fieldsValues = [bytes.fromhex(hexval) for hexval in self._comparator.lookupValues4FieldName(fieldname)]
        if len(fieldsValues) > 0:
            fieldLengths = Counter(len(bv) for bv in fieldsValues)
            # should normally be a constant value for this kind of fields
            mostCommonLen = fieldLengths.most_common(1)[0][0]
            logging.getLogger(__name__).debug(f"Field lengths of {fieldname}: {repr(fieldLengths)}")
            entropy = drv.entropy(intsFromNgrams(fieldsValues, self._endianness)) / (mostCommonLen * 8)
        else:
            entropy = numpy.nan
        return len(fieldsValues), entropy

    def typeAndLenEntropies(self):
        """
        Collect MSGtype/MSGlen true fields according to GroundTruth.fieldtypes[MSGtype.typelabel/MSGlen.typelabel]

        :return: list of lists of "field name", "type label", "sample count", and "entropy"
        """
        entropyList = list()
        for typelabel in [MSGtype.typelabel, MSGlen.typelabel]:
            for fieldname in GroundTruth.fieldtypes[typelabel]:
                # for each field name calculate entropy
                entropyList.append([
                    fieldname,
                    typelabel,
                    *self.entropyPerField(fieldname)
                ])
        return entropyList
